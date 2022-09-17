
from time import sleep
import os

# Initiate Selenium

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
 
# s=Service('C:/Users/Lenovo/Downloads/chromedriver.exe')
#driver = webdriver.Chrome(service=s)

print('Initializing Selenium driver\n')
options = webdriver.ChromeOptions()
user_data_dir = "%s\\chrome_profile"%(os.getcwd())
options.add_argument(r"user-data-dir=%s"%user_data_dir)

url='https://www.gst.gov.in/'
wd = webdriver.Chrome(ChromeDriverManager().install(), options=options)

wd.get(url)
sleep(2.0)


# ## Login

# In[7]:
username = 'Empliance_2709'
password = 'Rahul@123'

wd.get('https://www.gst.gov.in/')
login = wd.find_element_by_xpath("/html/body/div[1]/header/div[2]/div/div/ul/li").text
if login == "Login":
    # Login 
    print('login..')
    wd.get("https://services.gst.gov.in/services/login")
    sleep(2.0)

    wd.find_element_by_id('username').send_keys(username)
    wd.find_element_by_id('user_pass').send_keys(password)

    captcha = input("Enter Captcha to login : ")
    wd.find_element_by_id("captcha").send_keys(captcha)
    
    wd.find_element_by_css_selector('button.btn').click()
    
    # Go to Search by PAN
    sleep(5.0)    
    try:
        wd.find_elements_by_css_selector("a.btn.btn-primary")[-1].click()
    except: pass
    wd.find_elements_by_css_selector("li.dropdown.drpdwn.menuList")[2].click()
    wd.find_element_by_link_text('Search by PAN').click()
else: 
    pass



# In[8]:


from datetime import datetime
# Due Dates for GSTR3B
Dict1 = {'July': '20/8/2022',
        'June': '20/7/2022',
        'May': '20/06/2022',
        'April': '24/05/2022',
        'March': '20/04/2022', 
        'February': '20/03/2022',
        'January': '20/02/2022', 
        'December': '20/01/2022',
        'November': '20/12/2021',
        'October': '20/11/2021',
        'September': '20/10/2021',
        'August': '20/9/2022'}

# Due Dates for GSTR-1/IFF
Dict2= {'July': '11/8/2022',
        'June': '11/7/2022',
        'May': '11/06/2022',
        'April': '11/05/2022',
        'March': '11/04/2022', 
        'February': '11/03/2022',
        'January': '11/02/2022', 
        'December': '11/01/2022',
        'November': '11/12/2021',
        'October': '11/11/2021',
        'September': '11/10/2021',
        'August': '11/9/2022'}

# Due Dates for GSTR9/GSTR9C
Dict3= {'2020-2021': '12/02/2022', 
        '2019-2020': '31/03/2021',
        '2018-2019': '31/12/2021', 
        '2017-2018': '12/02/2021'}
# 23-12-2021 2020  2019-2020

# Due Dates for GSTR1
# Dict4 = {}


# ## Search by PAN

# In[9]:


k=0
#change PAN number
# pan='AACCG9224F'
pan = input("Enter PAN : ")

# wd.get('https://services.gst.gov.in/services/searchtpbypan')
# time.sleep(2.0)
Pan_search=wd.find_element_by_xpath('//*[@id="for_gstin"]')
Pan_search.clear()
Pan_search.send_keys(pan)
Pan_search.send_keys(Keys.RETURN)

submit=wd.find_element_by_xpath('//*[@id="lotsearch"]')
sleep(4)

tables=[]
print("Number of pages")
x=input()
p=int(x)+2
for i in range (0,int(x)):
    pan_table=wd.find_element_by_xpath("/html/body/div[2]/div[2]/div/div[2]/div/div/form/div[2]/div[2]/div/div/table/tbody").text
    tables.append(pan_table)
    sleep(2.0)
    
    #next page
    try: wd.find_element_by_xpath('/html/body/div[2]/div[2]/div/div[2]/div/div[1]/form/div[5]/div[3]/div/div/div/div/div/ul/li['+str(p)+']/a').click()
    except: pass
    


# ## GSTIN 

# In[87]:


from openpyxl.styles import Font
from openpyxl import Workbook

wb = Workbook()
wb.create_sheet('Sheet 1')
wb.create_sheet('Delayed')
wb.remove_sheet(wb['Sheet'])

sheet = wb['Sheet 1']
sheet2 = wb['Delayed']

GSTS=[]
i=3

sheet.cell(1, 1, 'Search Result based on PAN :'+str(pan))
sheet2.cell(1, 1,"Table of Delayed GSTs" )
sheet.cell(1, 1).font = Font(bold=True)

sheet.append(['S. No.', 'GSTIN/UIN', 'GSTIN/UIN Status','State'])
sheet2.append(['Table of','GST NO','Financial Year', 'Tax Period','Date of filing','Status','Empliance Check'])
    

for elem in tables :
    rows=elem.split('\n')
    
    for elem in rows:
        row=elem.split(' ')
        
        sheet.cell(i, 1, row[0])
        sheet.cell(i, 2, row[1])
        sheet.cell(i, 3, row[2])
        
        GSTS.append(row[1])    
        try:
            sheet.cell(i, 4, ' '.join(row[3:]))
        except:
            sheet.cell(i, 4, "-")
        i += 1
     

print(GSTS)
print("Total no of GSTS are :",len(GSTS))


# # Scraping

# In[88]:


def extract_info1():
    # Search Result based on GSTIN/UIN  
    res = wd.find_element_by_xpath('//*[@id="lottable"]/div[1]/div[1]/h4').text
    append_bold(sheet, res)
    
    table1 = wd.find_element_by_id("partners").find_element_by_class_name("tbl-format")
    t_rows= table1.find_elements_by_class_name("row")


    hd1 = t_rows[0].text.split('\n')
    for k in range(len(hd1)):
        if hd1[k] == 'Date of Registration':
            print(hd1[k+1])
            sheet.append(["Date of Registration", hd1[k+1]])

    hd2 = t_rows[1].text.split('\n')
    for k in range(len(hd2)):
        if hd2[k] == 'GSTIN / UIN Status':
            print(hd2[k+1])
            sheet.append(["GSTIN / UIN Status", hd2[k+1]])


# In[89]:


def extract_info2():
    # Name of the Proprietor / Director(s) / Promoter(s)
    append_bold(sheet, 'Name of the Proprietor / Director(s) / Promoter(s)')
    proprietors = wd.find_element_by_xpath('//*[@id="collapseOne"]/div').text.split('\n')

    for p in range(len(proprietors)):
        sheet.append([p+1, proprietors[p]])

    # Nature of Business Activities
    business = wd.find_element_by_id('collapseTwo').text.split('\n')
    append_bold(sheet, 'Nature of Business Activities')

    for p in range(len(business)):
        sheet.append([p+1, business[p]])

    # Nature Of Core Business Activity
    core_business = wd.find_element_by_id('ntcrbs').text.split('\n')
    append_bold(sheet, 'Nature Of Core Business Activity')

    for p in range(len(core_business)):
        sheet.append([p+1, core_business[p]])
    


# In[90]:


from datetime import datetime

def extract_filing_table():
    # Filing Table
    xpath_ = '//*[@id="partners"]/div[7]/div[2]/div[%s]/div/table/tbody'

    for i in range(1, 5):
        print(i)
        try:
            hd_x = wd.find_element_by_xpath('//*[@id="partners"]/div[7]/div[2]/div[%s]/div/h4'%i).text
        except: return
        print(hd_x)
        append_bold(sheet, hd_x)
        
        gstr = hd_x.split(' ')[-1]
        print(gstr)
        
        sheet.append(['Table of','GST NO','Financial Year', 'Tax Period','Date of filing','Status','Empliance Check'])

        for _ in wd.find_element_by_xpath(xpath_%(i)).text.split('\n'):
            row = _.split(' ')
            if gstr == 'GSTR3B':
                v_date = datetime.strptime(Dict1[row[1]], '%d/%m/%Y')
                
            elif gstr == 'GSTR-1/IFF':
                v_date = datetime.strptime(Dict2[row[1]], '%d/%m/%Y')
            
            elif (gstr == 'GSTR9') or (gstr == 'GSTR9C'):
                v_date = datetime.strptime(Dict3[row[0]], '%d/%m/%Y')
            # elif gstr == 'GSTR1':
            #     v_date = datetime.strptime(Dict4[row[1]], '%d/%m/%Y')
            else:
                d = [gstr, GST, *row]
                print(d)
                sheet.append(d)
                continue
            
            t_date = datetime.strptime(row[2], '%d/%m/%Y')
            
            if v_date < t_date:
                row.append("Delayed")
                sheet2.append([gstr, GST, *row])
            else:
                row.append("On Time")
                
            # print(row)
            d = [gstr, GST, *row]
            print(d)
            sheet.append(d)


# extract_filing_table()
# wb.save('my_gst.xlsx')


# In[91]:


def extract_liability():
    append_bold(sheet, "Liability table")

    xpth_l = '//*[@id="partners"]/div[7]/div[2]/div/div/div[%s]/table'
    for i in [2,3]:
        print(i)
        sheet.append([])
        try:
            hh = wd.find_element_by_xpath('//*[@id="partners"]/div[7]/div[2]/div/div/div[%s]/h4'%i).text
        except: continue
    
        sheet.append([hh])
        sheet.append(["Financial Year", "Tax Period", "% of Liability paid"])

        liability = wd.find_element_by_xpath(xpth_l%(i)).text.split('\n')[1:]
        for lb in liability:
            print(lb.split(' '))
            sheet.append(lb.split(' '))



# In[92]:


# Place of Business
def extract_place_of_business():
    append_bold(sheet, "Place of Business")
    sheet.append(["Type","Nature of Business Activities being carried out at Place of Business","Address"])
    
    wd.find_element_by_link_text('Place of Business').click()
    sleep(2.0)
    # address = wd.find_element_by_xpath('//*[@id="auth"]/div/div/div/table/tbody/tr/td[3]').text
    
    place_ = wd.find_element_by_xpath('//*[@id="auth"]/div/div/div/table/tbody')
    for tr in place_.find_elements_by_tag_name('tr'):
        d = []
        for _ in tr.find_elements_by_tag_name('td'):
            d.append(_.text)
        print(d)
        sheet.append(d)
        
    wd.find_element_by_link_text('Profile').click()
    sleep(2.0)
    
# extract_place_of_business()


# In[93]:


def extract_goods_services():
    # Dealing In Goods and Services
    goods_service = wd.find_element_by_xpath('//*[@id="partners"]/div[2]/div/div/div[1]/table/tbody')
    
    append_bold(sheet, 'Dealing In Goods and Services')
    m = sheet.max_row
    sheet.cell(m+1, 1, 'Goods')
    sheet.cell(m+1, 1).font = Font(bold=True)
    
    sheet.cell(m+1, 3, 'Services')
    sheet.cell(m+1, 3).font = Font(bold=True)

    for tr1 in goods_service.find_elements_by_tag_name('tr'):
        d = []
        for cc in tr1.find_elements_by_tag_name('td'):
            d.append(cc.text)
        print(d)
        sheet.append(d)
        
# extract_goods_services()


# In[95]:


def append_bold(sheet, text, s=2):
    m = sheet.max_row
    sheet.cell(m+s, 1, text)
    sheet.cell(m+s, 1).font = Font(bold=True)

# Go to search by GSTIN
wd.find_elements_by_css_selector("li.dropdown.drpdwn.menuList")[2].click()
wd.find_element_by_link_text('Search by GSTIN/UIN').click()
sleep(2.5)

for GST in GSTS[:]:
    print('GSTIN', GST)
    f1 = wd.find_element_by_id("for_gstin")
    wd.execute_script("return arguments[0].scrollIntoView();", f1)
    f1.clear()
    f1.send_keys(GST)
    wd.find_element_by_css_selector('button.btn').click()
    sleep(4.0)
    
    
    # Legal Name of Bussiness
    sheet.cell(1,2,"Legal Name of Business")
    sheet.cell(1,2).font = Font(bold=True)
    legal_name=wd.find_element_by_xpath('//*[@id="lottable"]/div[2]/div[1]/div').text.split('\n')[1]
    sheet.cell(1,3,legal_name)
    
    # EXTRACTING INFO
    extract_info1()
    extract_info2()
    
    # Click on Filing Table
    Filing_table=wd.find_element_by_xpath('//*[@id="filingTable"]')
    wd.execute_script("return arguments[0].scrollIntoView();", Filing_table)
    Filing_table.click()
    sleep(2.0) 

    extract_filing_table()
    
    # click on liability paid percentage
    liabity_element = wd.find_element_by_id('liabilitypaidpercentage')
    wd.execute_script("return arguments[0].scrollIntoView();", liabity_element)
    liabity_element.click()
    sleep(2.0)
    
    extract_liability()
    extract_place_of_business()
    extract_goods_services()
    


# In[96]:


text1 = [
    [
        "Liability paid percentage - Calculation methodology",
        "Liability paid percentage = (Liability paid / Liability auto drafted) *100",
        "This also includes the amount paid by the taxpayer for any period in Form DRC-03 by selecting the 'Cause of payment' as Liability mismatch - GSTR-1 to GSTR-3B.",
        "Liability Auto-drafted: This is the sum of total liability which is auto-drafted in Form GSTR-3B for a particular period from GSTR-1/IFF and GSTR-2B.",
        "For taxpayers opting to file return on monthly frequency, the liability paid percentage is computed for each period and for taxpayer opting to file return on quarterly frequency, the liability paid percentage is computed for the quarter."
    ],
    
    [ 
        "In case, liability auto drafted is Zero, then '-' is displayed",
        "In case, liability paid is Zero, then '0' is displayed",
        "In case both liabilities paid, and liability auto drafted are Zero, then '0' is displayed",
        "In case either the liability paid, or liability auto drafted is negative then:",
        "In case liability paid is equal to or more than the liability auto drafted then '100' is displayed",
        "In case liability paid is less than the liability auto drafted then '0' is displayed",
        "In case the liability paid percentage is less than 100, then the value is highlighted in red."
    ]
        ]

append_bold(sheet, text1[0][0])
for u in text1[0][1:]:
    sheet.append([u])
    
append_bold(sheet, "Note:")
for u in text1[1]:
    sheet.append([u])


# In[97]:

wb.save('gst.xlsx')
